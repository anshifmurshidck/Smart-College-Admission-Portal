import os
import datetime
from io import BytesIO
from flask import Blueprint, request, send_file, jsonify, make_response
from backend.db import db
from backend.middlewares.auth import token_required

# ReportLab libraries for PDF creation
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Openpyxl for Excel creation
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/generate', methods=['GET'])
@token_required()
def generate_report(current_user):
    report_type = request.args.get('type', 'admission') # 'admission', 'student', 'department'
    file_format = request.args.get('format', 'pdf') # 'pdf', 'excel'

    if report_type not in ('admission', 'student', 'department'):
        return jsonify({'message': 'Invalid report type requested'}), 400
    if file_format not in ('pdf', 'excel'):
        return jsonify({'message': 'Invalid file format requested'}), 400

    try:
        if report_type == 'admission':
            data = db.execute_read(
                """SELECT a.id, a.full_name, a.email, a.status, a.created_at, d.code as department_code 
                   FROM applications a
                   JOIN departments d ON a.department_id = d.id
                   ORDER BY a.created_at DESC"""
            )
            headers = ['Application ID', 'Full Name', 'Email', 'Department', 'Status', 'Applied Date']
            title = "Thought Minds Engineering College - Admission Applications Report"
            sheet_name = "Applications"
        elif report_type == 'student':
            data = db.execute_read(
                """SELECT s.id, s.full_name, s.email, s.enroll_date, d.code as department_code 
                   FROM students s
                   JOIN departments d ON s.department_id = d.id
                   ORDER BY s.enroll_date DESC"""
            )
            headers = ['Student ID', 'Full Name', 'Email', 'Department', 'Enroll Date']
            title = "Thought Minds Engineering College - Enrolled Students Database Report"
            sheet_name = "Students"
        else: # department
            data = db.execute_read(
                """SELECT d.code, d.name, d.head_of_department,
                   (SELECT COUNT(*) FROM applications WHERE department_id = d.id) as total_apps,
                   (SELECT COUNT(*) FROM students WHERE department_id = d.id) as enrolled_students
                   FROM departments d"""
            )
            headers = ['Code', 'Department Name', 'HOD', 'Total Applications', 'Enrolled Students']
            title = "Thought Minds Engineering College - Academic Departments Performance Report"
            sheet_name = "Departments"

        if file_format == 'pdf':
            return make_pdf_response(title, headers, data, report_type)
        else:
            return make_excel_response(title, headers, data, report_type, sheet_name)

    except Exception as e:
        print(f"[REPORT GENERATION ERROR]: {e}")
        return jsonify({'message': 'Error generating report', 'error': str(e)}), 500


def make_pdf_response(title, headers, data, report_type):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []

    # Palette
    NAVY = colors.HexColor("#0F172A")
    ROYAL = colors.HexColor("#2563EB")
    GOLD = colors.HexColor("#F59E0B")
    DARK_TEXT = colors.HexColor("#1E293B")
    LIGHT_BG = colors.HexColor("#F8FAFC")
    BORDER_COLOR = colors.HexColor("#E2E8F0")

    styles = getSampleStyleSheet()
    
    # Custom Typography Styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=NAVY,
        spaceAfter=6,
        alignment=1 # Centered
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=ROYAL,
        spaceAfter=15,
        alignment=1 # Centered
    )

    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=DARK_TEXT
    )

    cell_header_style = ParagraphStyle(
        'TableHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        textColor=colors.white
    )

    # Document Header
    story.append(Paragraph("THOUGHT MINDS ENGINEERING COLLEGE", title_style))
    story.append(Paragraph(title, subtitle_style))
    story.append(Spacer(1, 10))

    # Compile table data
    table_data = []
    # Headers
    table_data.append([Paragraph(h, cell_header_style) for h in headers])
    
    # Rows
    for row in data:
        row_cells = []
        if report_type == 'admission':
            row_cells.append(Paragraph(row['id'], cell_style))
            row_cells.append(Paragraph(row['full_name'], cell_style))
            row_cells.append(Paragraph(row['email'], cell_style))
            row_cells.append(Paragraph(row['department_code'], cell_style))
            
            # Status badge styling inside cell
            status = row['status']
            status_color = "#1E293B"
            if status == 'Approved':
                status_color = "#059669"
            elif status == 'Rejected':
                status_color = "#DC2626"
            elif status == 'Under Verification':
                status_color = "#D97706"
            
            status_style = ParagraphStyle(
                'StatusCell',
                parent=cell_style,
                fontName='Helvetica-Bold',
                textColor=colors.HexColor(status_color)
            )
            row_cells.append(Paragraph(status, status_style))
            
            date_val = row['created_at']
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)[:10]
            row_cells.append(Paragraph(date_str, cell_style))

        elif report_type == 'student':
            row_cells.append(Paragraph(row['id'], cell_style))
            row_cells.append(Paragraph(row['full_name'], cell_style))
            row_cells.append(Paragraph(row['email'], cell_style))
            row_cells.append(Paragraph(row['department_code'], cell_style))
            
            date_val = row['enroll_date']
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)[:10]
            row_cells.append(Paragraph(date_str, cell_style))

        else: # department
            row_cells.append(Paragraph(row['code'], cell_style))
            row_cells.append(Paragraph(row['name'], cell_style))
            row_cells.append(Paragraph(row['head_of_department'], cell_style))
            row_cells.append(Paragraph(str(row['total_apps']), cell_style))
            row_cells.append(Paragraph(str(row['enrolled_students']), cell_style))
        
        table_data.append(row_cells)

    # Determine column widths dynamically (Total available width = 540 pt)
    num_cols = len(headers)
    col_widths = [540 / num_cols] * num_cols
    # Adjust widths specifically for better sizing
    if report_type == 'admission':
        col_widths = [90, 110, 140, 70, 70, 60]
    elif report_type == 'student':
        col_widths = [100, 130, 160, 70, 80]
    elif report_type == 'department':
        col_widths = [50, 180, 110, 100, 100]

    # Create table
    t = Table(table_data, colWidths=col_widths)
    
    # Table Styling
    t_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), NAVY),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, BORDER_COLOR),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, LIGHT_BG]),
        ('TOPPADDING', (0,1), (-1,-1), 6),
        ('BOTTOMPADDING', (0,1), (-1,-1), 6),
    ])
    
    t.setStyle(t_style)
    story.append(t)

    # Build PDF document
    doc.build(story)
    
    buffer.seek(0)
    filename = f"{report_type}_report_{datetime.datetime.now().strftime('%Y%m%d%H%M')}.pdf"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )


def make_excel_response(title, headers, data, report_type, sheet_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    navy_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    light_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="0F172A")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Document Header
    ws.merge_cells("A1:F1")
    ws["A1"] = "THOUGHT MINDS ENGINEERING COLLEGE"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:F2")
    ws["A2"] = title
    ws["A2"].font = Font(name="Calibri", size=11, italic=True, color="2563EB")
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # Empty separator row
    ws.append([])

    # Add headers
    ws.append(headers)
    header_row_idx = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border

    # Add data rows
    row_idx = 5
    for row in data:
        row_vals = []
        if report_type == 'admission':
            date_val = row['created_at']
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)[:10]
            row_vals = [
                row['id'],
                row['full_name'],
                row['email'],
                row['department_code'],
                row['status'],
                date_str
            ]
        elif report_type == 'student':
            date_val = row['enroll_date']
            date_str = date_val.strftime('%Y-%m-%d') if hasattr(date_val, 'strftime') else str(date_val)[:10]
            row_vals = [
                row['id'],
                row['full_name'],
                row['email'],
                row['department_code'],
                date_str
            ]
        else: # department
            row_vals = [
                row['code'],
                row['name'],
                row['head_of_department'],
                row['total_apps'],
                row['enrolled_students']
            ]
        
        ws.append(row_vals)
        
        # Style row cells
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill
            
            # Apply color to status specifically
            if report_type == 'admission' and col_idx == 5:
                status = row['status']
                if status == 'Approved':
                    cell.font = Font(name="Calibri", size=11, bold=True, color="059669")
                elif status == 'Rejected':
                    cell.font = Font(name="Calibri", size=11, bold=True, color="DC2626")
                elif status == 'Under Verification':
                    cell.font = Font(name="Calibri", size=11, bold=True, color="D97706")
                    
        row_idx += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        # Avoid checking merged cell length extensively
        for cell in col[3:]: # Skip first 3 rows which contain headers
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{report_type}_report_{datetime.datetime.now().strftime('%Y%m%d%H%M')}.xlsx"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
